"""
Gap filler for Country_Data.xlsx.

Sources:
  PWT  -> pwt110 (2).xlsx       (same original source as Country_Data,
                                 values inserted without formatting)
  WDI  -> Iran/Qatar P_Data_Extract...   (supplementary columns, marked red)

Workflow:
  1) Iran-specific: fill delta + (Real internal rate of return) from PWT
  2) labsh (labor income share) as a NEW column for every country that
     has it in PWT (IRN, QAT, USA, SAU, KWT, IRQ, VEN; ARE/DZA stay empty)
  3) WDI supplementary columns (red):
       - "WDI: Inflation CPI (annual %)"   for Iran + Qatar
       - "WDI: Real interest rate (%)"     for Iran + Qatar
  4) Pre-computed regional aggregates as new "country" rows:
       - OPC = OPEC Core 6 (SAU, ARE, KWT, IRQ, DZA, VEN), no IRN/QAT
       - MEA = Middle East (BHR, IRQ, ISR, JOR, KWT, LBN, OMN, SAU, SYR,
               ARE, YEM, TUR, PSE), no IRN/QAT
       - WLD = World  (all PWT countries except IRN, QAT)
     Method: quantities (Y, K, Pop, N, ...) summed; rates (delta, labsh,
     pl_*, csh_*, hc) population-weighted; avh workforce-weighted. PWT
     outliers in price columns are masked before aggregation.
  5) Write report file with remaining gaps

Keeps the original Excel structure intact (openpyxl rather than a pandas
roundtrip) so user formatting is preserved.
"""

from pathlib import Path
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font

BASE = Path(__file__).resolve().parent
CD_FILE = BASE / "Country_Data.xlsx"
PWT_FILE = BASE / "pwt110.xlsx"
WDI_IRN = BASE / "Iran P_Data_Extract_From_World_Development_Indicators.xlsx"
WDI_QAT = BASE / "Qatar P_Data_Extract_From_World_Development_Indicators.xlsx"
REPORT_FILE = BASE / "data_gaps_report.md"

SUBJECT_ISOS = ["IRN", "QAT"]   # Excluded from all aggregates

AGGREGATE_DEFINITIONS = {
    # (iso_code, country_label, isos_or_None_for_World)
    ("OPC", "OPEC Core 6 (aggregate)"): [
        "SAU", "ARE", "KWT", "IRQ", "DZA", "VEN",
    ],
    ("MEA", "Middle East (aggregate)"): [
        "BHR", "IRQ", "ISR", "JOR", "KWT", "LBN", "OMN",
        "SAU", "SYR", "ARE", "YEM", "TUR", "PSE",
    ],
    ("WLD", "World (aggregate, ex IRN+QAT)"): None,  # all countries
}

# PWT short code -> Country_Data long column header
PWT_TO_CD = {
    "rgdpe":   "Inside Real GDP (Chained Prices$)",
    "rgdpo":   "Outside Real GDP (Chained Prices$)",
    "pop":     "Population",
    "emp":     "Workforce",
    "avh":     "avg. annual hours per worker",
    "hc":      "Human Capital Index",
    "ccon":    "Real Consuption (Current 2021 Prices$)",
    "cda":     "Domestic Absorbtion (Current 2021 Prices$)",
    "cgdpe":   "Inside Real GDP (Current 2021 Prices$)",
    "cgdpo":   "Outside Real GDP (Current 2021 Prices$)",
    "cn":      "Capital Stock (Current 2021 Prices$)",
    "ck":      "Capital Service level (Current 2021 Prices$)",
    "ctfp":    "TFP level (Current 2021 Prices$)",
    "cwtfp":   "Welfare level (Current 2021 Prices$)",
    "rgdpna":  "Real GDP  (Constant 2021 Prices$)",
    "rconna":  "Real Consuption (Constant 2021 Prices$)",
    "rdana":   "Domestic Absorbtion (Constant 2021 Prices$)",
    "rnna":    "Capital Stock (Constant 2021 Prices$)",
    "rkna":    "Capital Service level (Constant 2021 Prices$)",
    "rtfpna":  "TFP level (Constant 2021 Prices$)",
    "rwtfpna": "Welfare level (Constant 2021 Prices$)",
    "delta":   "delta",
    "pl_con":  "pl_con",
    "pl_da":   "pl_da",
    "pl_gdpo": "pl_gdpo",
    "csh_c":   "csh_c", "csh_i": "csh_i", "csh_g": "csh_g",
    "csh_x":   "csh_x", "csh_m": "csh_m", "csh_r": "csh_r",
    "pl_c":    "pl_c", "pl_i": "pl_i", "pl_g": "pl_g",
    "pl_x":    "pl_x", "pl_m": "pl_m", "pl_n": "pl_n", "pl_k": "pl_k",
    # labsh handled separately because it has no long name in original CD
}

# Quantity columns – summed across member countries
SUM_PWT_COLS = [
    "rgdpe", "rgdpo", "pop", "emp", "ccon", "cda", "cgdpe", "cgdpo",
    "cn", "ck", "rgdpna", "rconna", "rdana", "rnna", "rkna",
]

# Rate-like columns – population-weighted mean across member countries
POP_WEIGHTED_PWT_COLS = [
    "delta", "hc", "labsh",
    "csh_c", "csh_i", "csh_g", "csh_x", "csh_m", "csh_r",
    "pl_con", "pl_da", "pl_gdpo",
    "pl_c", "pl_i", "pl_g", "pl_x", "pl_m", "pl_n", "pl_k",
]

# Workforce-weighted mean
EMP_WEIGHTED_PWT_COLS = ["avh"]

# Index/level columns that should NOT be summed (would be meaningless) —
# these are skipped for aggregates.
SKIPPED_PWT_COLS = ["ctfp", "cwtfp", "rtfpna", "rwtfpna", "irr", "xr"]


# ---------------------------------------------------------------------------
# 1) Load PWT and index requested series
# ---------------------------------------------------------------------------
def load_pwt() -> pd.DataFrame:
    pwt = pd.read_excel(PWT_FILE, sheet_name="Data")
    pwt = pwt.dropna(subset=["countrycode", "year"])
    pwt["year"] = pwt["year"].astype(int)
    return pwt


# ---------------------------------------------------------------------------
# 2) Load WDI -> long format
# ---------------------------------------------------------------------------
def load_wdi(file: Path, iso: str) -> pd.DataFrame:
    df = pd.read_excel(file, sheet_name="Data")
    df = df.dropna(subset=["Series Name"])
    year_cols = [c for c in df.columns if "[YR" in str(c)]
    long = df.melt(
        id_vars=["Series Name", "Series Code"],
        value_vars=year_cols,
        var_name="year_str",
        value_name="value",
    )
    long["year"] = long["year_str"].str.extract(r"(\d{4})").astype(int)
    long["value"] = pd.to_numeric(long["value"], errors="coerce")
    long["countrycode"] = iso
    return long.drop(columns=["year_str"])


def get_wdi_series(wdi_long: pd.DataFrame, name: str) -> pd.Series:
    s = wdi_long.loc[wdi_long["Series Name"] == name, ["year", "value"]]
    return s.set_index("year")["value"]


# ---------------------------------------------------------------------------
# 3) Edit workbook (openpyxl, preserves original structure)
# ---------------------------------------------------------------------------
RED_FONT = Font(color="FFFF0000")  # ARGB red


def find_header_row(ws):
    """Return (row_index, {header_name: column_index}) for sheet 'Data'."""
    for r in range(1, 5):
        cells = [(c, ws.cell(row=r, column=c).value) for c in range(1, ws.max_column + 1)]
        names = [v for _, v in cells if isinstance(v, str)]
        if "countrycode" in names and "year" in names:
            mapping = {v: c for c, v in cells if isinstance(v, str)}
            return r, mapping
    raise RuntimeError("Header row not found")


def ensure_column(ws, header_row: int, col_map: dict, header: str) -> int:
    if header in col_map:
        return col_map[header]
    new_col = ws.max_column + 1
    ws.cell(row=header_row, column=new_col, value=header)
    col_map[header] = new_col
    return new_col


def fill_country_data():
    pwt = load_pwt()
    wdi_irn = load_wdi(WDI_IRN, "IRN")
    wdi_qat = load_wdi(WDI_QAT, "QAT")

    # WDI series we want to add
    wdi_cpi_irn = get_wdi_series(wdi_irn, "Inflation, consumer prices (annual %)")
    wdi_cpi_qat = get_wdi_series(wdi_qat, "Inflation, consumer prices (annual %)")
    wdi_rir_irn = get_wdi_series(wdi_irn, "Real interest rate (%)")
    wdi_rir_qat = get_wdi_series(wdi_qat, "Real interest rate (%)")

    wb = load_workbook(CD_FILE)
    ws = wb["Data"]
    header_row, col_map = find_header_row(ws)

    # Index: (countrycode, year) -> row index
    iso_col = col_map["countrycode"]
    yr_col = col_map["year"]
    row_index: dict[tuple[str, int], int] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        iso = ws.cell(row=r, column=iso_col).value
        yr = ws.cell(row=r, column=yr_col).value
        if iso and yr is not None:
            try:
                row_index[(str(iso), int(yr))] = r
            except (TypeError, ValueError):
                pass

    fills = {"pwt_count": 0, "wdi_count": 0, "skipped": 0}
    log = []

    # ------ PWT gap-filling ------
    pwt_jobs = [
        # (Country_Data column header, PWT column, [target ISOs])
        ("delta", "delta", ["IRN"]),
        ("(Real internal rate of return)", "irr", ["IRN"]),
    ]
    for cd_col, pwt_col, isos in pwt_jobs:
        if cd_col not in col_map:
            log.append(f"  ! Column '{cd_col}' not in Country_Data – skipped")
            continue
        c = col_map[cd_col]
        for iso in isos:
            pwt_sub = pwt[pwt["countrycode"] == iso].set_index("year")[pwt_col]
            for yr, v in pwt_sub.dropna().items():
                key = (iso, int(yr))
                if key not in row_index:
                    continue
                r = row_index[key]
                if ws.cell(row=r, column=c).value in (None, ""):
                    ws.cell(row=r, column=c, value=float(v))
                    fills["pwt_count"] += 1

    # labsh as new column
    labsh_col = ensure_column(ws, header_row, col_map, "labsh")
    for iso in ["IRN", "QAT", "USA", "SAU", "KWT", "IRQ", "VEN"]:
        pwt_sub = pwt[pwt["countrycode"] == iso].set_index("year")["labsh"]
        for yr, v in pwt_sub.dropna().items():
            key = (iso, int(yr))
            if key not in row_index:
                continue
            r = row_index[key]
            if ws.cell(row=r, column=labsh_col).value in (None, ""):
                ws.cell(row=r, column=labsh_col, value=float(v))
                fills["pwt_count"] += 1

    # ------ WDI gap-filling (red) ------
    wdi_cpi_col = ensure_column(ws, header_row, col_map, "WDI: Inflation CPI (annual %)")
    wdi_rir_col = ensure_column(ws, header_row, col_map, "WDI: Real interest rate (%)")

    for iso, cpi, rir in [("IRN", wdi_cpi_irn, wdi_rir_irn),
                          ("QAT", wdi_cpi_qat, wdi_rir_qat)]:
        for yr, v in cpi.dropna().items():
            key = (iso, int(yr))
            if key in row_index:
                r = row_index[key]
                cell = ws.cell(row=r, column=wdi_cpi_col, value=float(v))
                cell.font = RED_FONT
                fills["wdi_count"] += 1
        for yr, v in rir.dropna().items():
            key = (iso, int(yr))
            if key in row_index:
                r = row_index[key]
                cell = ws.cell(row=r, column=wdi_rir_col, value=float(v))
                cell.font = RED_FONT
                fills["wdi_count"] += 1

    # Color WDI column headers red
    for col in (wdi_cpi_col, wdi_rir_col):
        ws.cell(row=header_row, column=col).font = RED_FONT

    wb.save(CD_FILE)
    return fills, log


# ---------------------------------------------------------------------------
# 4) Compute and write regional aggregates as new "country" rows
# ---------------------------------------------------------------------------
def _mask_outliers(df: pd.DataFrame) -> pd.DataFrame:
    """Mask price-level outliers (i_outlier='Outlier') in price columns."""
    if "i_outlier" not in df.columns:
        return df
    out_mask = df["i_outlier"].astype(str).str.lower().str.contains("outlier")
    if not out_mask.any():
        return df
    df = df.copy()
    price_cols = ["pl_con", "pl_da", "pl_gdpo",
                  "pl_c", "pl_i", "pl_g", "pl_x", "pl_m", "pl_n", "pl_k",
                  "cgdpe", "cgdpo"]
    for c in price_cols:
        if c in df.columns:
            df.loc[out_mask, c] = np.nan
    return df


def compute_aggregate(pwt: pd.DataFrame, isos: list[str] | None
                      ) -> pd.DataFrame:
    """
    Build a year-indexed DataFrame for the aggregate.
    isos=None  -> World (all countries except SUBJECT_ISOS).
    Returns DataFrame with PWT short-name columns.
    """
    if isos is None:
        member_isos = [c for c in pwt["countrycode"].unique()
                       if c not in SUBJECT_ISOS]
    else:
        member_isos = [c for c in isos if c not in SUBJECT_ISOS]

    sub = pwt[pwt["countrycode"].isin(member_isos)].copy()
    sub = _mask_outliers(sub)

    # Year filter: require at least half of the members to have rgdpna and pop
    n_required = max(2, len(member_isos) // 2)
    cov = sub.groupby("year")[["rgdpna", "pop"]].count()
    valid_years = cov[(cov["rgdpna"] >= n_required)
                      & (cov["pop"] >= n_required)].index

    # Sum quantities
    sum_cols = [c for c in SUM_PWT_COLS if c in sub.columns]
    sums = (
        sub.groupby("year")[sum_cols]
        .sum(min_count=1)
        .loc[valid_years]
    )

    # Pop-weighted means
    def _wmean(g: pd.DataFrame, col: str, weight_col: str = "pop") -> float:
        if col not in g.columns or weight_col not in g.columns:
            return np.nan
        w = g[weight_col]
        v = g[col]
        m = v.notna() & w.notna()
        if m.sum() == 0:
            return np.nan
        return float((v[m] * w[m]).sum() / w[m].sum())

    weighted = pd.DataFrame(index=valid_years)
    for col in POP_WEIGHTED_PWT_COLS:
        if col not in sub.columns:
            continue
        s = sub.groupby("year").apply(
            lambda g, c=col: _wmean(g, c, "pop"),
            include_groups=False,
        )
        weighted[col] = s.loc[valid_years]

    for col in EMP_WEIGHTED_PWT_COLS:
        if col not in sub.columns:
            continue
        s = sub.groupby("year").apply(
            lambda g, c=col: _wmean(g, c, "emp"),
            include_groups=False,
        )
        weighted[col] = s.loc[valid_years]

    return pd.concat([sums, weighted], axis=1)


def _find_last_data_row(ws, iso_col: int, header_row: int) -> int:
    """Return the row index of the last non-empty row in `iso_col`.
    Used instead of ws.max_row, which is inflated by formatted-but-empty
    cells in this particular workbook (openpyxl reports max_row=~13686
    although real data ends around row 681)."""
    last = header_row
    for r in range(header_row + 1, ws.max_row + 1):
        v = ws.cell(row=r, column=iso_col).value
        if v is not None and v != "":
            last = r
    return last


def _trim_trailing_empty_rows(ws, iso_col: int, header_row: int) -> int:
    """Delete every fully-empty row after the last data row to keep the
    workbook clean. Returns the number of rows removed."""
    last_real = _find_last_data_row(ws, iso_col, header_row)
    if ws.max_row > last_real:
        n = ws.max_row - last_real
        ws.delete_rows(last_real + 1, n)
        return n
    return 0


def write_aggregates_to_country_data() -> dict:
    """
    Compute OPC / MEA / WLD aggregates from PWT and write them as new
    "country" rows into Country_Data.xlsx. Existing rows with these
    countrycodes are removed and rewritten so the operation is idempotent.
    """
    pwt = load_pwt()
    wb = load_workbook(CD_FILE)
    ws = wb["Data"]
    header_row, col_map = find_header_row(ws)

    # Ensure 'labsh' header exists (created in earlier step normally)
    ensure_column(ws, header_row, col_map, "labsh")

    iso_col = col_map["countrycode"]

    # 1) Drop any existing rows for these aggregate ISOs (idempotency).
    #    Iterate only over rows that actually have content.
    aggregate_isos = {iso for (iso, _name), _ in AGGREGATE_DEFINITIONS.items()}
    last_data_row = _find_last_data_row(ws, iso_col, header_row)
    rows_to_delete = []
    for r in range(header_row + 1, last_data_row + 1):
        v = ws.cell(row=r, column=iso_col).value
        if isinstance(v, str) and v in aggregate_isos:
            rows_to_delete.append(r)
    for r in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(r, 1)

    # 2) Trim trailing empty (but formatted) rows so new rows append
    #    directly after the real data instead of after a long blank gap.
    trimmed = _trim_trailing_empty_rows(ws, iso_col, header_row)

    counts = {}
    # Track the next free row explicitly – do not trust ws.max_row.
    next_row = _find_last_data_row(ws, iso_col, header_row) + 1

    # 3) Append fresh aggregate rows
    for (iso, label), isos in AGGREGATE_DEFINITIONS.items():
        agg = compute_aggregate(pwt, isos)
        n_added = 0
        for year, row in agg.iterrows():
            ws.cell(row=next_row, column=col_map["countrycode"], value=iso)
            ws.cell(row=next_row, column=col_map["country"], value=label)
            ws.cell(row=next_row, column=col_map["currency_unit"], value="—")
            ws.cell(row=next_row, column=col_map["year"], value=int(year))

            for pwt_col, val in row.items():
                if pd.isna(val):
                    continue
                cd_col_name = "labsh" if pwt_col == "labsh" else PWT_TO_CD.get(pwt_col)
                if cd_col_name is None or cd_col_name not in col_map:
                    continue
                ws.cell(row=next_row, column=col_map[cd_col_name],
                        value=float(val))
            next_row += 1
            n_added += 1
        counts[iso] = n_added

    wb.save(CD_FILE)
    counts["_trimmed_empty_rows"] = trimmed
    return counts


# ---------------------------------------------------------------------------
# 5) Write report of remaining gaps
# ---------------------------------------------------------------------------
def build_report():
    cd = pd.read_excel(CD_FILE, sheet_name="Data")
    cd = cd.dropna(subset=["countrycode", "year"]).copy()
    cd["year"] = cd["year"].astype(int)

    isos = ["IRN", "QAT", "USA", "SAU", "ARE", "KWT", "IRQ", "DZA", "VEN"]
    cols_check = [
        ("delta", "Depreciation rate (Solow)"),
        ("(Real internal rate of return)", "Real interest rate / Rental rate"),
        ("labsh", "Labor share (for α calibration)"),
        ("avg. annual hours per worker", "Hours/worker (for N = emp×hours)"),
        ("TFP level (Constant 2021 Prices$)", "TFP (Constant)"),
        ("Capital Stock (Constant 2021 Prices$)", "Capital Stock (K)"),
        ("Capital Service level (Constant 2021 Prices$)", "Capital Services"),
        ("Welfare level (Constant 2021 Prices$)", "Welfare"),
        ("csh_c", "Consumption share"),
        ("csh_i", "Investment share"),
        ("Population", "Population"),
        ("Workforce", "Workforce (emp)"),
        ("Real GDP  (Constant 2021 Prices$)", "Real GDP"),
    ]

    lines = ["# Country_Data.xlsx – Data Gaps Report", ""]
    lines.append(f"Generated: {pd.Timestamp.now():%Y-%m-%d %H:%M}")
    lines.append("")
    lines.append("## Remaining gaps after gap-filling from PWT 11.0 and WDI")
    lines.append("")
    lines.append("Display: `non-null / total` per country × variable. Empty cells shown as `–`.")
    lines.append("")

    header = ["Variable"] + isos
    lines.append("| " + " | ".join(header) + " |")
    lines.append("| " + " | ".join(["---"] * len(header)) + " |")

    for col, label in cols_check:
        if col not in cd.columns:
            row = [f"**{label}**"] + ["MISSING"] * len(isos)
            lines.append("| " + " | ".join(row) + " |")
            continue
        row = [f"**{label}**"]
        for iso in isos:
            sub = cd[cd["countrycode"] == iso]
            n = sub[col].notna().sum()
            tot = len(sub)
            if n == 0:
                cell = "– (0)"
            elif n == tot:
                cell = f"OK ({n})"
            else:
                cell = f"{n}/{tot}"
            row.append(cell)
        lines.append("| " + " | ".join(row) + " |")

    lines.append("")
    lines.append("## Data sources")
    lines.append("")
    lines.append("- **Black values**: from PWT 11.0 (the original source of Country_Data)")
    lines.append("- **<span style=\"color:red\">Red values</span>**: added from World Development Indicators (Iran & Qatar)")
    lines.append("  - Column `WDI: Inflation CPI (annual %)`")
    lines.append("  - Column `WDI: Real interest rate (%)`")
    lines.append("")
    lines.append("## Known unresolvable gaps")
    lines.append("")
    lines.append("- `avg. annual hours per worker` for MENA countries: only available 2005–2023 (PWT's own limitation).")
    lines.append("  Workaround: use N = Workforce without hours correction.")
    lines.append("- `labsh`, `irr`, `TFP`, `Capital Service`, `Welfare` for **ARE** and **DZA**: not available in PWT 11.0.")
    lines.append("  Workaround for Solow aggregation: exclude ARE/DZA from calculations that require these quantities, or")
    lines.append("  assume α = 1/3 as a textbook value and compute TFP yourself.")
    lines.append("- `TFP/Capital Service/Welfare` 2020–2023 for several countries: PWT 11.0 computes these with a lag.")
    lines.append("- `csh_c` for Iraq only 12/54: 1990s sanctions period, acceptable gap (Iraq only used in OPEC aggregate).")

    REPORT_FILE.write_text("\n".join(lines), encoding="utf-8")


# ---------------------------------------------------------------------------
def main():
    fills, log = fill_country_data()
    print(f"PWT values inserted: {fills['pwt_count']}")
    print(f"WDI values inserted (red marked): {fills['wdi_count']}")
    if log:
        print("\nLog:")
        for l in log:
            print(l)

    print("\nWriting regional aggregates to Country_Data.xlsx ...")
    counts = write_aggregates_to_country_data()
    for iso, n in counts.items():
        print(f"  {iso}: {n} year-rows added")

    build_report()
    print(f"\nReport: {REPORT_FILE}")


if __name__ == "__main__":
    main()
